import csv
import pandas as pd
import requests
from requests_kerberos import HTTPKerberosAuth, OPTIONAL
import openpyxl
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows

goog_url = "https://monitorportal.amazon.com/mws?Action=GetGraph&Version=2007-07-07&SchemaName1=Search&Pattern1" \
           "=marketplace%3D%24DSM5%24%20methodname%3D%24ALL%24%20metric%3DDSM5-SLAM%20%28LabelScan%20OR%20LabelZone" \
           "%20OR%20PackageCleared%20OR%20NoRead%20OR%20MultiRead%29&Period1=OneHour&Stat1=sum&SchemaName2=Search" \
           "&Pattern2=marketplace%3D%24DSM5%24%20methodname%3D%24processPackageVerified%24%20metric%3DDSM5%20Shipped" \
           "%20&Stat2=n&HeightInPixels=250&WidthInPixels=600&GraphTitle=SLAM%20REPORT%20v5&TZ=CST6CDT@TZ%3A%20CST6CDT" \
           "&StartTime1=-PT12H&EndTime1=-PT0H&OutputFormat=CSV_TRANSPOSE "


def download_slam_data(csv_url):
    download = requests.get(csv_url, auth=HTTPKerberosAuth(OPTIONAL), verify=False)

    decoded_csv = download.content.decode('utf-8')

    cr = csv.reader(decoded_csv.splitlines(), delimiter=',')
    my_list = list(cr)
    for row in my_list:
        print(my_list)
    df = pd.DataFrame(my_list)
    df.to_csv("SLAM_file.csv")


download_slam_data(goog_url)

df = pd.read_csv('SLAM_file.csv', header=None, index_col=None)
wb = load_workbook(filename='C:\Users\bryaburg\Desktop\Python Project\Slam Report\SLAM_Report.xlsm')
ws = wb.create_sheet()
ws.title = 'ToCopy'

for r in dataframe_to_rows(df, index=False, header=False):
    ws.append(r)

wb.save('C:\Users\bryaburg\Desktop\Python Project\Slam Report\SLAM_file.xlsm')





