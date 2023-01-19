import csv
import pandas as pd
import requests
from requests_kerberos import HTTPKerberosAuth, OPTIONAL
import openpyxl
import io

goog_url = "https://monitorportal.amazon.com/mws?Action=GetGraph&Version=2007-07-07&SchemaName1=Search&Pattern1" \
           "=marketplace%3D%24DSM5%24%20methodname%3D%24ALL%24%20metric%3DDSM5-SLAM%20%28LabelScan%20OR%20LabelZone" \
           "%20OR%20PackageCleared%20OR%20NoRead%20OR%20MultiRead%29&Period1=OneHour&Stat1=sum&SchemaName2=Search" \
           "&Pattern2=marketplace%3D%24DSM5%24%20methodname%3D%24processPackageVerified%24%20metric%3DDSM5%20Shipped" \
           "%20&Stat2=n&HeightInPixels=250&WidthInPixels=600&GraphTitle=SLAM%20REPORT%20v5&TZ=CST6CDT@TZ%3A%20CST6CDT" \
           "&StartTime1=-PT12H&EndTime1=-PT0H&OutputFormat=CSV_TRANSPOSE "

# This grabs the slam data and saves it in a CSV file
def download_slam_data(csv_url):
    
    #This Auth me getting the info from the URL
    download = requests.get(csv_url, auth=HTTPKerberosAuth(OPTIONAL), verify=False)
    
    # This decodes it and then saves it to a CSV file
    decoded_csv = download.content.decode('utf-8')
    cr = csv.reader(decoded_csv.splitlines(), delimiter=',')
    my_list = list(cr)
    df = pd.DataFrame(my_list)
    df.to_csv("SLAM_file.csv", header=False, index=False)
    
    # Open the existing Excel file and put in csv data
    with pd.ExcelWriter('C:/Users/bryaburg/Desktop/Python_Projects/Python_Projects/Slam Report/SLAM Report.xlsx', mode='a', engine= "openpyxl", if_sheet_exists = 'replace') as writer:
        # Read the excel file
        df = pd.read_csv('SLAM_file.csv')
        #Convert data in Dataframe to float
        for col in df.columns:
            try:
                df[col] = df[col].astype(float)
            except ValueError:
                pass
        # Write the data to the specified sheet
        df.to_excel(writer,sheet_name='DATA', startrow=0, startcol=0,  index=False, header=False)
        
        # Load the existing excel file
        book = openpyxl.load_workbook('C:/Users/bryaburg/Desktop/Python_Projects/Python_Projects/Slam Report/SLAM Report.xlsx')

        # Get the sheet 'DATA'
        data_sheet = book['DATA']

        # Get the number of rows and columns in the sheet
        num_rows = data_sheet.max_row
        num_cols = data_sheet.max_column

        # Iterate over the rows and columns of the sheet 'DATA'
        for i in range(1, num_rows+1):
            for j in range(1, num_cols+1):
                # Get the cell
                cell = data_sheet.cell(row=i, column=j)

                # Try to convert the cell value to a number
                try:
                    cell.value = float(cell.value)  #float() argument must be a string or a real number, not 'NoneType'
                except ValueError:
                    pass

        # Save the changes to the excel file
        book.save('C:/Users/bryaburg/Desktop/Python_Projects/Python_Projects/Slam Report/SLAM Report.xlsx')

download_slam_data(goog_url)